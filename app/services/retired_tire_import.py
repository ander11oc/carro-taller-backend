from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.entities import (
    FleetOperationCompany,
    RetiredTireRecord,
    TireBrandDesign,
    TireRetirementCondition,
)


DATA_SHEET = "Data"
CONDITIONS_SHEET = "BASE DE CONDICIONES"
BRANDS_SHEET = "MARCA_LLANTAS"
COMPANIES_SHEET = "Operacion_Empresas"


@dataclass(frozen=True)
class RetiredTireImportResult:
    retired_tires: int
    conditions: int
    brand_designs: int
    operation_companies: int


def import_retired_tire_workbook(
    db: Session, workbook_path: str, tenant_id: str
) -> RetiredTireImportResult:
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        _replace_tenant_data(db, tenant_id)
        conditions = _import_conditions(db, workbook, tenant_id)
        brand_designs = _import_brand_designs(db, workbook, tenant_id)
        operation_companies = _import_operation_companies(db, workbook, tenant_id)
        retired_tires = _import_retired_tires(db, workbook, tenant_id)
        db.commit()
        return RetiredTireImportResult(
            retired_tires=retired_tires,
            conditions=conditions,
            brand_designs=brand_designs,
            operation_companies=operation_companies,
        )
    except Exception:
        db.rollback()
        raise
    finally:
        workbook.close()


def _replace_tenant_data(db: Session, tenant_id: str) -> None:
    for model in (
        RetiredTireRecord,
        TireRetirementCondition,
        TireBrandDesign,
        FleetOperationCompany,
    ):
        db.query(model).filter(model.tenant_id == tenant_id).delete()


def _import_retired_tires(db: Session, workbook, tenant_id: str) -> int:
    rows = _iter_sheet_records(workbook, DATA_SHEET)
    count = 0
    for source_row, row in rows:
        if not any(row.values()):
            continue
        if not _value(row, "Empresa") and not _value(row, "#Interno"):
            continue
        db.add(
            RetiredTireRecord(
                tenant_id=tenant_id,
                source_sheet=DATA_SHEET,
                source_row=source_row,
                quantity=_int_value(row, "Cantidad"),
                mount_date=_date_value(row, "FechaMontaje"),
                dismount_date=_date_value(row, "FechaDeontaje"),
                month=_str_value(row, "Mes"),
                year=_int_value(row, "Año"),
                company=_str_value(row, "Empresa"),
                typology=_str_value(row, "Tipologia"),
                design=_str_value(row, "Diseño"),
                brand=_str_value(row, "Marca"),
                dimension=_str_value(row, "Dimension"),
                ply_rating=_str_value(row, "PlyRating"),
                internal_number=_str_value(row, "#Interno"),
                observations=_str_value(row, "Observaciones"),
                repair=_str_value(row, "Reparacion"),
                repair_status=_str_value(row, "EstadoReparaciones"),
                patch_count=_int_value(row, "CantidaddeParches"),
                condition_code=_str_value(row, "Codigo-Prodes"),
                tire_area=_str_value(row, "AreaLlanta"),
                retirement_condition=_str_value(
                    row, "CondicionesdeRetiroEvidenciadas"
                ),
                original_tread_depth=_float_value(row, "ProfundidadOriginal"),
                exterior_tread=_float_value(row, "Exterior"),
                center_tread=_float_value(row, "Centro"),
                interior_tread=_float_value(row, "Interior"),
                min_tread=_float_value(row, "Prof.Min"),
                max_tread=_float_value(row, "Prof.Max"),
                tread_diff=_float_value(row, "DifProf.BDR"),
                unused_bdr_pct=_float_value(row, "%BDRsinutilizar"),
                tread_wear=_str_value(row, "DesgasteBDR"),
                new_or_retread=_str_value(row, "Nueva/Reencacuhe"),
                retread_band_design=_str_value(row, "DiseñoBandaReen"),
                application=_str_value(row, "APLICACIÓN"),
                lives=_int_value(row, "#Vidas"),
                casing_use=_float_value(row, "UsoCarcasa"),
                unused_mm=_float_value(row, "Mmsinutilizar"),
                new_tire_value=_float_value(row, "ValorllantaNueva"),
                retread_value=_float_value(row, "ValorllantaReencauche"),
                total_cost=_float_value(row, "CostoTotal"),
                work_time_years=_float_value(row, "Tiempodetrabajo(año)"),
                final_new_tire_km=_float_value(row, "KmfinalLlantaNueva"),
                final_retread_1_km=_float_value(row, "Kmfinal 1Reen"),
                final_retread_2_km=_float_value(row, "Kmfinal 2Reen"),
                final_retread_3_km=_float_value(row, "Kmfinal 3Reen"),
                regravation_km=_float_value(row, "KmRegrabacion"),
                total_km=_float_value(row, "KmRecorrido(total)"),
                cpk=_float_value(row, "Cpk"),
            )
        )
        count += 1
    return count


def _import_conditions(db: Session, workbook, tenant_id: str) -> int:
    count = 0
    for _, row in _iter_sheet_records(workbook, CONDITIONS_SHEET):
        if not any(row.values()):
            continue
        db.add(
            TireRetirementCondition(
                tenant_id=tenant_id,
                code_description=_str_value(row, "CODIGO_DESCRIPCION"),
                description=_str_value(row, "Descripción"),
                column_code=_str_value(row, "Columna1"),
                zone=_str_value(row, "ZONA"),
                motive_group=_str_value(row, "Grupo Motivos"),
                area_code=_int_value(row, "Codigo_Area"),
            )
        )
        count += 1
    return count


def _import_brand_designs(db: Session, workbook, tenant_id: str) -> int:
    count = 0
    for _, row in _iter_sheet_records(workbook, BRANDS_SHEET):
        if not any(row.values()):
            continue
        db.add(
            TireBrandDesign(
                tenant_id=tenant_id,
                design=_str_value(row, "DISEÑO"),
                brand=_str_value(row, "MARCA"),
                nks=_int_value(row, "NKS"),
                application=_str_value(row, "APLICACIÓN"),
            )
        )
        count += 1
    return count


def _import_operation_companies(db: Session, workbook, tenant_id: str) -> int:
    count = 0
    for _, row in _iter_sheet_records(workbook, COMPANIES_SHEET):
        if not any(row.values()):
            continue
        db.add(
            FleetOperationCompany(
                tenant_id=tenant_id,
                company=_str_value(row, "Empresa"),
                operation=_str_value(row, "Operación"),
                route=_str_value(row, "Ruta"),
            )
        )
        count += 1
    return count


def _iter_sheet_records(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        return
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return
    normalized_headers = [_normalize_header(header) for header in headers]
    for offset, values in enumerate(rows, start=2):
        record = {
            header: value
            for header, value in zip(normalized_headers, values)
            if header
        }
        yield offset, record


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split())


def _value(row: dict[str, Any], key: str) -> Any:
    return row.get(_normalize_header(key))


def _str_value(row: dict[str, Any], key: str) -> str:
    value = _value(row, key)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _int_value(row: dict[str, Any], key: str) -> int | None:
    value = _value(row, key)
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _float_value(row: dict[str, Any], key: str) -> float | None:
    value = _value(row, key)
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _date_value(row: dict[str, Any], key: str) -> date | None:
    value = _value(row, key)
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None
