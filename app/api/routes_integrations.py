import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any, Callable

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.api.permissions import require_module_action
from app.db.session import get_db
from app.models.entities import (
    FuelLog,
    IntegrationEvent,
    IntegrationRun,
    MaintenanceOrder,
    MediaAsset,
    NotificationMessage,
    PurchaseRequest,
    Tire,
    TireEvent,
    Vehicle,
    VehicleTirePosition,
)
from app.schemas.integrations import IntegrationEventOut, IntegrationRunOut, IntegrationWebhookRequest


router = APIRouter(prefix="/integrations", tags=["integrations"])

TIRE_MECHANICAL_TERMS = ("eje", "ejes", "suspension", "freno", "frenos", "alineacion")


def _tenant(user: dict) -> str:
    return user["tenant_id"]


def _actor(user: dict) -> str:
    return user.get("email", "")


def _as_float(value: Any, default: float | None = None) -> float | None:
    if value is None or value == "":
        return default
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return default


def _as_int(value: Any, default: int = 0) -> int:
    number = _as_float(value)
    return default if number is None else int(number)


def _as_date(value: Any) -> date:
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str) and value.strip():
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
        except ValueError:
            try:
                return date.fromisoformat(value[:10])
            except ValueError:
                return date.today()
    return date.today()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _cell_value(value: Any) -> str | int | float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    return text if text else None


def _rows_to_records(rows: list[list[Any]], first_row_header: bool, columns: list[str] | None = None) -> list[dict[str, Any]]:
    non_empty_rows = [row for row in rows if any(_cell_value(cell) is not None for cell in row)]
    if not non_empty_rows:
        return []

    if first_row_header:
        headers = [str(_cell_value(cell) or f"column_{index + 1}").strip() for index, cell in enumerate(non_empty_rows[0])]
        data_rows = non_empty_rows[1:]
    else:
        headers = columns or [f"column_{index + 1}" for index in range(len(non_empty_rows[0]))]
        data_rows = non_empty_rows

    return [
        {
            headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}": _cell_value(cell)
            for index, cell in enumerate(row)
        }
        for row in data_rows
    ]


def parse_integration_file(
    filename: str,
    content: bytes,
    sheet_name: str | None = None,
    first_row_header: bool = True,
    columns: list[str] | None = None,
) -> list[dict[str, Any]]:
    lower_name = filename.lower()
    if lower_name.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        selected_sheet = sheet_name if sheet_name and sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        rows = [list(row) for row in workbook[selected_sheet].iter_rows(values_only=True)]
        return _rows_to_records(rows, first_row_header, columns)

    text = content.decode("utf-8-sig")
    sample = text[:2048]
    delimiter = "\t" if "\t" in sample else ";" if sample.count(";") > sample.count(",") else ","
    rows = [row for row in csv.reader(StringIO(text), delimiter=delimiter)]
    return _rows_to_records(rows, first_row_header, columns)


def _find_vehicle(db: Session, user: dict, plate: Any) -> Vehicle | None:
    plate_text = _text(plate)
    if not plate_text:
        return None
    return (
        db.query(Vehicle)
        .filter(Vehicle.tenant_id == _tenant(user), Vehicle.plate == plate_text)
        .first()
    )


def _find_tire(db: Session, user: dict, serial: Any) -> Tire | None:
    serial_text = _text(serial)
    if not serial_text:
        return None
    return (
        db.query(Tire)
        .filter(Tire.tenant_id == _tenant(user), Tire.serial_number == serial_text)
        .first()
    )


def _integration_event(
    db: Session,
    run: IntegrationRun,
    user: dict,
    event_type: str,
    status_value: str,
    message: str,
    payload: dict[str, Any],
    entity_type: str = "",
    entity_id: int | None = None,
) -> None:
    db.add(
        IntegrationEvent(
            tenant_id=_tenant(user),
            run_id=run.id,
            system=run.system,
            event_type=event_type,
            entity_type=entity_type,
            entity_id=entity_id,
            status=status_value,
            message=message,
            payload=payload,
        )
    )


def _process_gps(record: dict[str, Any], db: Session, run: IntegrationRun, user: dict) -> None:
    vehicle = _find_vehicle(db, user, record.get("plate"))
    if not vehicle:
        raise ValueError(f"Vehiculo no encontrado para placa {record.get('plate')}")

    odometer = _as_float(record.get("odometer", record.get("mileage")))
    speed = _as_float(record.get("speed"), 0) or 0

    if odometer is not None:
        if odometer >= vehicle.mileage:
            vehicle.mileage = odometer
        else:
            _integration_event(
                db,
                run,
                user,
                "gps_mileage_rejected",
                "warning",
                f"Kilometraje menor al anterior para {vehicle.plate}: {odometer} < {vehicle.mileage}",
                record,
                "vehicle",
                vehicle.id,
            )

    if speed >= 85:
        _integration_event(
            db,
            run,
            user,
            "gps_speed_critical",
            "warning",
            f"Velocidad critica para {vehicle.plate}: {speed:g} km/h",
            record,
            "vehicle",
            vehicle.id,
        )


def _process_fuel(record: dict[str, Any], db: Session, run: IntegrationRun, user: dict) -> None:
    vehicle = _find_vehicle(db, user, record.get("plate"))
    if not vehicle:
        raise ValueError(f"Vehiculo no encontrado para placa {record.get('plate')}")

    logged_on = _as_date(record.get("date") or record.get("logged_on"))
    station = _text(record.get("station"))
    mileage = _as_float(record.get("mileage"), vehicle.mileage) or vehicle.mileage
    existing = (
        db.query(FuelLog)
        .filter(
            FuelLog.tenant_id == _tenant(user),
            FuelLog.vehicle_id == vehicle.id,
            FuelLog.logged_on == logged_on,
            FuelLog.station == station,
            FuelLog.mileage == mileage,
        )
        .first()
    )
    if not existing:
        db.add(
            FuelLog(
                tenant_id=_tenant(user),
                vehicle_id=vehicle.id,
                liters=_as_float(record.get("liters", record.get("gallons")), 0) or 0,
                mileage=mileage,
                cost=_as_float(record.get("cost"), 0) or 0,
                station=station,
                logged_on=logged_on,
            )
        )
    if mileage > vehicle.mileage:
        vehicle.mileage = mileage


def _process_maintenance(record: dict[str, Any], db: Session, run: IntegrationRun, user: dict) -> None:
    vehicle = _find_vehicle(db, user, record.get("plate"))
    if not vehicle:
        raise ValueError(f"Vehiculo no encontrado para placa {record.get('plate')}")

    title = _text(record.get("title") or record.get("order") or "Orden externa")
    description = _text(record.get("description") or record.get("diagnosis"))
    order = MaintenanceOrder(
        tenant_id=_tenant(user),
        vehicle_id=vehicle.id,
        title=title,
        description=description,
        status=_text(record.get("status")) or "open",
        priority=_text(record.get("priority")) or "normal",
        scheduled_for=_as_date(record.get("date") or record.get("scheduled_for")),
        cost=_as_float(record.get("cost"), 0) or 0,
    )
    db.add(order)
    db.flush()

    lower = f"{title} {description}".lower()
    if any(term in lower for term in TIRE_MECHANICAL_TERMS):
        mounted = (
            db.query(VehicleTirePosition)
            .filter(
                VehicleTirePosition.tenant_id == _tenant(user),
                VehicleTirePosition.vehicle_id == vehicle.id,
                VehicleTirePosition.tire_id.isnot(None),
            )
            .all()
        )
        linked_positions = [
            {"tire_id": position.tire_id, "position": position.position_code}
            for position in mounted
            if position.tire_id
        ]
        if not linked_positions:
            linked_positions = [
                {"tire_id": tire.id, "position": tire.position}
                for tire in db.query(Tire)
                .filter(Tire.tenant_id == _tenant(user), Tire.vehicle_id == vehicle.id)
                .all()
            ]
        for position in linked_positions:
            db.add(
                TireEvent(
                    tenant_id=_tenant(user),
                    tire_id=position["tire_id"],
                    vehicle_id=vehicle.id,
                    event_type="maintenance_linked",
                    event_date=_as_date(record.get("date")),
                    position=position["position"],
                    novelty=f"{title}. {description}".strip(),
                    created_by=_actor(user),
                    created_role=user.get("role", ""),
                )
            )


def _process_purchases(record: dict[str, Any], db: Session, run: IntegrationRun, user: dict) -> None:
    external_id = _text(record.get("external_id") or record.get("order_number"))
    query = db.query(PurchaseRequest).filter(PurchaseRequest.tenant_id == _tenant(user))
    if external_id:
        exists = query.filter(PurchaseRequest.external_id == external_id).first()
    else:
        exists = query.filter(
            PurchaseRequest.request_type == (_text(record.get("type")) or "tire"),
            PurchaseRequest.origin == _text(record.get("origin")),
            PurchaseRequest.provider_suggested == _text(record.get("provider")),
            PurchaseRequest.status == (_text(record.get("status")) or "suggested"),
        ).first()
    if exists:
        return

    db.add(
        PurchaseRequest(
            tenant_id=_tenant(user),
            request_type=_text(record.get("type")) or "tire",
            origin=_text(record.get("origin")),
            provider_suggested=_text(record.get("provider")),
            quantity=max(_as_int(record.get("quantity"), 1), 1),
            priority=_text(record.get("priority")) or "normal",
            status=_text(record.get("status")) or "suggested",
            source_system=run.system,
            external_id=external_id,
            notes=_text(record.get("notes")),
        )
    )


def _process_retread(record: dict[str, Any], db: Session, run: IntegrationRun, user: dict) -> None:
    tire = _find_tire(db, user, record.get("serial") or record.get("serial_number"))
    if not tire:
        raise ValueError(f"Llanta no encontrada para serial {record.get('serial')}")

    status_value = _text(record.get("status")).lower()
    event_type = {
        "sent": "retread_sent",
        "received": "retread_received",
        "rejected": "retread_rejected",
    }.get(status_value, "retread_update")

    if event_type == "retread_received":
        tire.status = "available"
        tire.life_cycle = "retread"
        tire.retread_band = _text(record.get("band")) or tire.retread_band
        tire.provider = _text(record.get("provider")) or tire.provider
    elif event_type == "retread_rejected":
        tire.status = "rejected"
    elif event_type == "retread_sent":
        tire.status = "retread"

    db.add(
        TireEvent(
            tenant_id=_tenant(user),
            tire_id=tire.id,
            vehicle_id=tire.vehicle_id,
            event_type=event_type,
            event_date=_as_date(record.get("date")),
            position=tire.position,
            provider=_text(record.get("provider")),
            cost=_as_float(record.get("cost")),
            novelty=_text(record.get("concept") or record.get("result") or record.get("status")),
            evidence_url=_text(record.get("evidence_url")),
            created_by=_actor(user),
            created_role=user.get("role", ""),
        )
    )


def _process_erp(record: dict[str, Any], db: Session, run: IntegrationRun, user: dict) -> None:
    serial = record.get("serial") or record.get("serial_number")
    tire = _find_tire(db, user, serial) if serial else None
    _integration_event(
        db,
        run,
        user,
        "erp_cost_received",
        "processed",
        f"Costo/factura ERP recibida {record.get('invoice', '')}".strip(),
        record,
        "tire" if tire else "erp",
        tire.id if tire else None,
    )


def _process_media_storage(record: dict[str, Any], db: Session, run: IntegrationRun, user: dict) -> None:
    filename = _text(record.get("filename") or record.get("name") or "evidencia")
    entity_type = _text(record.get("entity_type"))
    entity_id = record.get("entity_id")
    entity_id_int = _as_int(entity_id, 0) if entity_id not in (None, "") else None
    url = _text(record.get("url") or record.get("public_url") or record.get("evidence_url")) or f"local://media/{filename}"
    existing = (
        db.query(MediaAsset)
        .filter(
            MediaAsset.tenant_id == _tenant(user),
            MediaAsset.url == url,
            MediaAsset.entity_type == entity_type,
            MediaAsset.entity_id == entity_id_int,
        )
        .first()
    )
    if existing:
        return
    db.add(
        MediaAsset(
            tenant_id=_tenant(user),
            url=url,
            filename=filename,
            content_type=_text(record.get("content_type")) or "application/octet-stream",
            entity_type=entity_type,
            entity_id=entity_id_int,
            evidence_type=_text(record.get("evidence_type")) or "evidence",
            uploaded_by=_actor(user),
            status=_text(record.get("status")) or "available",
            source=run.system,
        )
    )


def _process_notifications(record: dict[str, Any], db: Session, run: IntegrationRun, user: dict) -> None:
    channel = _text(record.get("channel")) or "email"
    template = _text(record.get("template")) or "alerta_operativa"
    entity_type = _text(record.get("entity_type"))
    entity_id = record.get("entity_id")
    entity_id_int = _as_int(entity_id, 0) if entity_id not in (None, "") else None
    existing = (
        db.query(NotificationMessage)
        .filter(
            NotificationMessage.tenant_id == _tenant(user),
            NotificationMessage.channel == channel,
            NotificationMessage.template == template,
            NotificationMessage.entity_type == entity_type,
            NotificationMessage.entity_id == entity_id_int,
            NotificationMessage.status.in_(["pending", "sent"]),
        )
        .first()
    )
    if existing:
        return
    db.add(
        NotificationMessage(
            tenant_id=_tenant(user),
            channel=channel,
            template=template,
            entity_type=entity_type,
            entity_id=entity_id_int,
            recipient=_text(record.get("recipient")),
            status=_text(record.get("status")) or "pending",
            payload=record,
        )
    )


PROCESSORS: dict[str, Callable[[dict[str, Any], Session, IntegrationRun, dict], None]] = {
    "gps_telematics": _process_gps,
    "erp_accounting": _process_erp,
    "maintenance_system": _process_maintenance,
    "fuel": _process_fuel,
    "purchases": _process_purchases,
    "retread_providers": _process_retread,
    "media_storage": _process_media_storage,
    "notifications": _process_notifications,
}


def _process_run(run: IntegrationRun, records: list[dict[str, Any]], db: Session, user: dict) -> IntegrationRun:
    processor = PROCESSORS.get(run.system)
    if not processor:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Integracion no soportada: {run.system}")

    run.status = "processing"
    run.started_at = datetime.utcnow()
    run.total_records = len(records)
    run.processed_records = 0
    run.failed_records = 0
    run.errors = []
    db.flush()

    errors: list[str] = []
    for index, record in enumerate(records, start=1):
        try:
            processor(record, db, run, user)
            run.processed_records += 1
        except Exception as exc:  # pragma: no cover - message is persisted for operators.
            run.failed_records += 1
            message = f"Fila/evento {index}: {exc}"
            errors.append(message)
            _integration_event(db, run, user, "record_failed", "failed", message, record)

    run.errors = errors
    run.finished_at = datetime.utcnow()
    run.status = "failed" if errors and run.processed_records == 0 else "completed_with_errors" if errors else "completed"
    db.commit()
    db.refresh(run)
    return run


@router.get("/runs", response_model=list[IntegrationRunOut])
def list_integration_runs(
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    require_module_action(user, "integrations", "read")
    return (
        db.query(IntegrationRun)
        .filter(IntegrationRun.tenant_id == _tenant(user))
        .order_by(IntegrationRun.created_at.desc(), IntegrationRun.id.desc())
        .limit(limit)
        .all()
    )


@router.get("/runs/{run_id}", response_model=dict)
def get_integration_run(run_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    require_module_action(user, "integrations", "read")
    run = (
        db.query(IntegrationRun)
        .filter(IntegrationRun.tenant_id == _tenant(user), IntegrationRun.id == run_id)
        .first()
    )
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration run not found")
    events = (
        db.query(IntegrationEvent)
        .filter(IntegrationEvent.tenant_id == _tenant(user), IntegrationEvent.run_id == run.id)
        .order_by(IntegrationEvent.id.asc())
        .all()
    )
    return {
        "run": IntegrationRunOut.model_validate(run),
        "events": [IntegrationEventOut.model_validate(event) for event in events],
    }


@router.post("/runs/{run_id}/retry", response_model=IntegrationRunOut)
def retry_integration_run(run_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    require_module_action(user, "integrations", "update")
    run = (
        db.query(IntegrationRun)
        .filter(IntegrationRun.tenant_id == _tenant(user), IntegrationRun.id == run_id)
        .first()
    )
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Integration run not found")
    payload = run.payload or {}
    records = payload.get("records") or []
    return _process_run(run, records, db, user)


@router.post("/{system}/webhook", response_model=IntegrationRunOut)
def create_integration_webhook(
    system: str,
    payload: IntegrationWebhookRequest,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    require_module_action(user, "integrations", "create")
    payload_data = payload.model_dump()
    run = IntegrationRun(
        tenant_id=_tenant(user),
        system=system,
        source=payload.source or "webhook",
        status="pending",
        total_records=len(payload.records),
        payload=payload_data,
        errors=[],
        created_by=_actor(user),
    )
    db.add(run)
    db.flush()
    return _process_run(run, payload.records, db, user)


@router.post("/{system}/upload", response_model=IntegrationRunOut)
async def upload_integration_records(
    system: str,
    file: UploadFile = File(...),
    sheet_name: str = Form(""),
    first_row_header: bool = Form(True),
    columns: str = Form(""),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    require_module_action(user, "integrations", "import")
    raw = await file.read()
    column_list = [item.strip() for item in columns.split(",") if item.strip()]
    records = parse_integration_file(
        file.filename or f"{system}.csv",
        raw,
        sheet_name=sheet_name or None,
        first_row_header=first_row_header,
        columns=column_list or None,
    )
    payload = IntegrationWebhookRequest(records=records, source="upload")
    return create_integration_webhook(system, payload, db, user)
