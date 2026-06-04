from datetime import datetime
import unittest

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db.base import Base
from app.models.entities import (
    FleetOperationCompany,
    RetiredTireRecord,
    TireBrandDesign,
    TireRetirementCondition,
)
from app.services.retired_tire_import import import_retired_tire_workbook
from app.services.retired_tire_summary import get_retired_tire_summary


def _make_workbook(path):
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(
        [
            "Cantidad",
            "FechaMontaje",
            "FechaDeontaje",
            "Mes",
            "Año",
            "Empresa",
            "Tipologia",
            "Diseño",
            "Marca",
            "Dimension",
            "PlyRating",
            "#Interno",
            "Observaciones",
            "Reparacion",
            "EstadoReparaciones",
            "CantidaddeParches",
            "Codigo-Prodes",
            "AreaLlanta",
            "CondicionesdeRetiroEvidenciadas",
            "ProfundidadOriginal",
            "Exterior",
            "Centro",
            "Interior",
            "Prof.Min",
            "Prof.Max",
            "DifProf.BDR",
            "%BDRsinutilizar",
            "DesgasteBDR",
            "Nueva/Reencacuhe",
            "DiseñoBandaReen",
            "APLICACIÓN",
            "#Vidas",
            "UsoCarcasa",
            "Mmsinutilizar",
            "ValorllantaNueva",
            "ValorllantaReencauche",
            "CostoTotal",
            "Tiempodetrabajo(año)",
            "KmfinalLlantaNueva",
            "Kmfinal\n1Reen",
            "Kmfinal\n2Reen",
            "Kmfinal\n3Reen",
            "KmRegrabacion",
            "KmRecorrido(total)",
            "Cpk",
        ]
    )
    data.append(
        [
            18,
            datetime(2018, 12, 4),
            datetime(2019, 1, 18),
            "Enero",
            2019,
            "CC CARGA",
            "Transporte",
            "VDL",
            "VIKRANT",
            "295/80R22.5",
            None,
            "88",
            "Retiro inicial",
            "SI",
            "OK",
            1,
            "F1",
            "PESTAÑA",
            "Pestaña cristalizada",
            22,
            20,
            20,
            20,
            20,
            20,
            0,
            0.818182,
            "DesgasteNormal",
            "R",
            "HDR2",
            "TRACCION",
            2,
            2.090909,
            18,
            1000,
            300,
            1300,
            1.2,
            10000,
            2000,
            0,
            0,
            0,
            12000,
            0.11,
        ]
    )

    conditions = workbook.create_sheet("BASE DE CONDICIONES")
    conditions.append(["CODIGO_DESCRIPCION", "Descripción", "Columna1", "ZONA", "Grupo Motivos", "Codigo_Area"])
    conditions.append(["F1", "Pestaña cristalizada", "F1", "PESTAÑA", "Operación", 2])

    brands = workbook.create_sheet("MARCA_LLANTAS")
    brands.append(["DISEÑO", "MARCA", "NKS", "APLICACIÓN"])
    brands.append(["VDL", "VIKRANT", 16, "TRACCION"])

    companies = workbook.create_sheet("Operacion_Empresas")
    companies.append(["Empresa", "Operación", "Ruta"])
    companies.append(["CC CARGA", "Transporte Carroceria", "Nacional"])

    workbook.save(path)


class RetiredTireImportTest(unittest.TestCase):
    def test_import_retired_tire_workbook_loads_records_and_catalogs(self):
        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "retired_tires.xlsx"
            _make_workbook(workbook_path)

            engine = create_engine("sqlite:///:memory:")
            Base.metadata.create_all(engine)
            SessionLocal = sessionmaker(bind=engine)
            db = SessionLocal()

            result = import_retired_tire_workbook(
                db, str(workbook_path), tenant_id="tenant_test"
            )

            self.assertEqual(result.retired_tires, 1)
            self.assertEqual(result.conditions, 1)
            self.assertEqual(result.brand_designs, 1)
            self.assertEqual(result.operation_companies, 1)

            record = db.query(RetiredTireRecord).one()
            self.assertEqual(record.tenant_id, "tenant_test")
            self.assertEqual(record.company, "CC CARGA")
            self.assertEqual(record.internal_number, "88")
            self.assertEqual(record.retirement_condition, "Pestaña cristalizada")
            self.assertEqual(record.final_retread_1_km, 2000)

            self.assertEqual(
                db.query(TireRetirementCondition).one().code_description, "F1"
            )
            self.assertEqual(db.query(TireBrandDesign).one().design, "VDL")
            self.assertEqual(db.query(FleetOperationCompany).one().route, "Nacional")
            db.close()
            engine.dispose()

    def test_get_retired_tire_summary_groups_imported_data(self):
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        SessionLocal = sessionmaker(bind=engine)
        db = SessionLocal()

        db.add_all(
            [
                RetiredTireRecord(
                    tenant_id="tenant_test",
                    source_row=2,
                    year=2020,
                    month="Enero",
                    company="CC CARGA",
                    brand="VIKRANT",
                    tire_area="PESTAÑA",
                    condition_code="F1",
                    retirement_condition="Pestaña cristalizada",
                    new_or_retread="R",
                    casing_use=2.0,
                    cpk=0.15,
                ),
                RetiredTireRecord(
                    tenant_id="tenant_test",
                    source_row=3,
                    year=2020,
                    month="Enero",
                    company="CC CARGA",
                    brand="KUMHO",
                    tire_area="COSTADO",
                    condition_code="D1",
                    retirement_condition="Herida costado",
                    new_or_retread="N",
                    casing_use=1.0,
                    cpk=0.25,
                ),
                RetiredTireRecord(
                    tenant_id="tenant_test",
                    source_row=4,
                    year=2020,
                    month="Febrero",
                    company="CC CARGA",
                    brand="VIKRANT",
                    tire_area="PESTAÑA",
                    condition_code="F1",
                    retirement_condition="Pestaña cristalizada",
                    new_or_retread="R",
                    casing_use=3.0,
                    cpk=0.35,
                ),
            ]
        )
        db.commit()

        summary = get_retired_tire_summary(db, "tenant_test")

        self.assertEqual(summary["total"], 3)
        self.assertEqual(summary["retread_count"], 2)
        self.assertEqual(summary["new_count"], 1)
        self.assertEqual(summary["avg_casing_use"], 2.0)
        self.assertEqual(summary["avg_cpk"], 0.25)
        self.assertEqual(summary["by_brand"][0], {"label": "VIKRANT", "value": 2})
        self.assertEqual(summary["by_area"][0], {"label": "PESTAÑA", "value": 2})
        self.assertEqual(
            summary["by_condition"][0],
            {"label": "Pestaña cristalizada", "value": 2},
        )
        self.assertEqual(summary["by_month"], [{"label": "2020 Enero", "value": 2}, {"label": "2020 Febrero", "value": 1}])

        filtered = get_retired_tire_summary(
            db,
            "tenant_test",
            brands=["KUMHO"],
        )
        self.assertEqual(filtered["total"], 1)
        self.assertEqual(filtered["by_brand"], [{"label": "KUMHO", "value": 1}])
        self.assertEqual(filtered["by_month"], [{"label": "2020 Enero", "value": 1}])

        db.close()
        engine.dispose()


if __name__ == "__main__":
    unittest.main()
